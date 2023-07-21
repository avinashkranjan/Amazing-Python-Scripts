from openpyxl import load_workbook
from openpyxl import Workbook
import os


# Read data from active worksheet and return it as a list
def reader(file):
    global path
    abs_file = os.path.join(path, file)
    wb_sheet = load_workbook(abs_file).active
    rows = []
    # min_row is set to 2, ignore the first row which contains headers
    for row in wb_sheet.iter_rows(min_row=2):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    return rows


# You can replace these with your own headers for the table
headers = ['Nume', 'Prenume', 'Titlu', 'Editura', 'Cota', 'Pret', 'An']
# Unified excel name
workbook_name = input('Unified Workbook name ')
book = Workbook()
sheet = book.active
# Specify path
path = input('Path: ')
# Get all files from folder
files = os.listdir(path)
for file in files:
    rows = reader(file)
    for row in rows:
        sheet.append(row)
    book.save(filename=workbook_name)
